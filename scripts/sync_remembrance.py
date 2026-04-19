"""Populate every character's `remembrance_slots` from the xlsx items sheet.

Before: only Undertaker had remembrance_slots (incomplete — 2 effect IDs + 1
text-only locked_label). Fallback logic in constraints.py meant every other
character pulled Undertaker's Glass Necklace as slot-1 Remembrance.

After: each of the 10 Nightfarers has its own final-chapter Remembrance relic
as slot-1 fixed effects, using the full 3-effect-ID list from the xlsx
compendium so the solver + UI can show concrete effect text everywhere.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path("/mnt/c/Users/dharm/Documents/nightreign/data/nightreign data.xlsx")
CHAR_DIR = ROOT / "data" / "characters"

# Primary remembrance relic per character — picked to be the "late-game"
# quest reward that players most commonly run. xlsx_id is unique.
PRIMARY: dict[str, int] = {
    "wylder":     11001,   # The Wylder's Earring
    "guardian":   12002,   # Witch's Brooch
    "ironeye":    13002,   # Edge of Order
    "duchess":    14002,   # Blessed Iron Coin
    "raider":     15002,   # Black Claw Necklace
    "revenant":   16002,   # Old Portrait
    "recluse":    17002,   # Bone-Like Stone
    "executor":   18002,   # Golden Sprout
    "undertaker": 19051,   # Glass Necklace
    "scholar":    19001,   # Note "My Dear Successor"
}

_COLOR_HINT = {
    "Red": "red", "Blue": "blue", "Green": "green", "Yellow": "yellow",
}


def _load_items(wb):
    ws = wb["items"]
    out: dict[int, dict] = {}
    # Header row varies; skip until we land on data (numeric id, string name).
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = row[0]
        if item_id is None or not row[4]:
            continue
        if not isinstance(item_id, (int, float)):
            continue
        attrs = [int(e) for e in (row[5], row[7], row[9]) if e is not None]
        out[int(item_id)] = {
            "name": str(row[4]),
            "color": row[2],
            "attrs": attrs,
            "location": row[11] or "",
        }
    return out


def main() -> int:
    wb = load_workbook(XLSX, data_only=True)
    items = _load_items(wb)

    updated = 0
    for char_file in sorted(CHAR_DIR.glob("*.json")):
        data = json.load(char_file.open("r", encoding="utf-8"))
        cid = data["id"]
        primary_id = PRIMARY.get(cid)
        if primary_id is None or primary_id not in items:
            print(f"skip {cid}: no primary mapping")
            continue
        relic = items[primary_id]
        if len(relic["attrs"]) < 1:
            print(f"skip {cid}: relic {primary_id} has no attrs")
            continue

        color_hint = _COLOR_HINT.get(relic["color"], "green")
        slot_entry = {
            "name": relic["name"],
            "source": f"xlsx_item_{primary_id}",
            "fixed_effects": relic["attrs"],
            "locked_labels": [],           # text-only labels no longer needed
            "color_hint": color_hint,
        }
        data["remembrance_slots"] = [slot_entry]

        char_file.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n",
                             encoding="utf-8")
        print(f"✓ {cid:11s}  slot1 ← {relic['name']!r}  effects={relic['attrs']}  hint={color_hint}")
        updated += 1

    print(f"\nUpdated {updated}/{len(list(CHAR_DIR.glob('*.json')))} character files.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
