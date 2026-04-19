"""One-shot import from the relics.pro compendium xlsx.

Regenerates (from the xlsx provided via --xlsx):

  data/effects_enriched.json   — effect_id → { category, effect_text, stack_*,
                                               is_dn, roll, parsed: {...} }
  data/named_relics.json       — 121 items with explicit E1/E2/E3 effect IDs
  data/buffs.json              — 87 dormant powers (buff id, template, values)
  data/bosses.json             — 162 boss encounters (id, name, type, threat)

The CE table remains authoritative for effect id / tier / flag / group / name.
This script layers the xlsx data on top — CE id ⇄ xlsx id is 100% identical,
so merging is by id lookup.

Usage:
  uv run python scripts/import_xlsx.py --xlsx /path/to/nightreign-data.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"


# ────────────────────────────────────────────────────────────────────────
# parsers — convert prose EFFECT text to machine-readable fields
# ────────────────────────────────────────────────────────────────────────

_PCT = r"(\d+(?:\.\d+)?)\s*%"

_MULT_PATTERNS: list[tuple[re.Pattern, str]] = [
    # Scoped damage boost ("Increases physical damage by 10.5%")
    (re.compile(rf"[Ii]ncreases?\s+(physical|magic|fire|lightning|holy)\s+damage\s+(?:by\s+)?{_PCT}"),
     "damage_mult"),
    # "Increases attack power by N%"
    (re.compile(rf"[Ii]ncreases?\s+attack\s+power\s+(?:by\s+)?{_PCT}"),
     "attack_power_mult"),
    # "Increases Hammer/Sword/etc damage by N%" — weapon-scoped
    (re.compile(rf"[Ii]ncreases?\s+(Hammer|Sword|Bow|Dagger|Spear|Axe|Halberd|Knife|Fists|Staff|Seal)\s+damage\s+(?:by\s+)?{_PCT}"),
     "weapon_damage_mult"),
    # Plain "Increases damage by N%" (no scope keyword, usually melee/all)
    (re.compile(rf"[Ii]ncreases?\s+damage\s+(?:by\s+)?{_PCT}"),
     "generic_damage_mult"),
    # Tiered values "5% / 6.5% / 9%" — used for +N variants; capture highest
    (re.compile(rf"{_PCT}\s*/\s*{_PCT}\s*/\s*{_PCT}"),
     "tiered_pct"),
    # "N% damage buff / damage increase"
    (re.compile(rf"{_PCT}\s+damage\s+(?:buff|increase)"),
     "generic_damage_mult"),
    (re.compile(rf"[Ll]owers?\s+(physical|magic|fire|lightning|holy)\s+damage\s+(?:by\s+)?{_PCT}"),
     "damage_reduction"),
    (re.compile(rf"[Mm]aximum\s+(HP|FP|[Ss]tamina)\s+raised\s+by\s+{_PCT}"),
     "max_stat"),
    (re.compile(rf"[Ii]ncreases?\s+(physical|magic|fire|lightning|holy)\s+damage\s+negation\s+by\s+{_PCT}"),
     "damage_negation"),
    (re.compile(rf"[Rr]estores?\s+{_PCT}\s+of\s+max\s+(HP|FP|[Ss]tamina)"),
     "restore_pct"),
    (re.compile(rf"[Ll]owers?\s+(FP|[Ss]tamina)\s+consumption\s+by\s+{_PCT}"),
     "consumption_reduction"),
    (re.compile(rf"[Cc]ritical\s+(?:hit\s+)?damage\s+(?:up|increased?)\s+(?:by\s+)?{_PCT}"),
     "crit_mult"),
    # Utility: gauge / charge bonuses
    (re.compile(rf"[Ii]ncreases?\s+(ultimate|art|skill|character\s+skill)\s+gauge\s+(?:gain|charge)\s+(?:by\s+)?{_PCT}"),
     "gauge_boost"),
    (re.compile(rf"[Ii]ncreases?\s+(Hammer|Sword|Bow|Dagger|Spear|Axe|Halberd|Knife|Fists|Staff|Seal)\s+damage\s+by\s+\+?{_PCT}"),
     "weapon_damage_mult"),
    # Utility: cooldown reduction (skill / art)
    (re.compile(rf"[Rr]educes?\s+(skill|art|ultimate)\s+cooldown\s+(?:by\s+)?{_PCT}"),
     "cooldown_reduction"),
]

_DURATION_RE = re.compile(r"[Ll]asts?\s+(?:for\s+)?(?:for\s+)?(\d+)\s+seconds?")
_TRIGGER_MAP = {
    "after taking a hit": "damage_taken",
    "when getting hit": "damage_taken",
    "after continuous attacks": "successive_attacks",
    "successive attacks": "successive_attacks",
    "critical hit": "critical_hit",
    "while art is active": "ult_active",
    "ultimate art is active": "ult_active",
    "while character skill is active": "skill_active",
    "character skill is active": "skill_active",
    "guard counter": "guard_counter",
}


def parse_effect_text(text: str) -> dict:
    """Best-effort regex extraction. Any prose we can't parse stays in effect_text."""
    if not text:
        return {}
    out: dict = {}
    low = text.lower()

    tiered = _MULT_PATTERNS[4][0].search(text) if len(_MULT_PATTERNS) > 4 else None
    # When text has "5% / 6.5% / 9%" tiered variants (+N relic tiers), only
    # the highest value matters for the solver since we display the +3 roll.
    highest_tier: float | None = None
    if tiered:
        try:
            highest_tier = max(float(x) for x in tiered.groups())
        except ValueError:
            highest_tier = None

    for pat, kind in _MULT_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        g = m.groups()
        if kind == "damage_mult":
            out["damage_scope"] = g[0].lower()
            out["damage_mult_pct"] = float(g[1])
        elif kind == "attack_power_mult":
            out["attack_power_mult_pct"] = highest_tier or float(g[0])
        elif kind == "weapon_damage_mult":
            out["weapon_scope"] = g[0].lower()
            out["weapon_damage_mult_pct"] = highest_tier or float(g[1])
        elif kind == "generic_damage_mult":
            out["generic_damage_mult_pct"] = highest_tier or float(g[0])
        elif kind == "tiered_pct":
            # already consumed — keep the max in case nothing else matched
            out.setdefault("generic_damage_mult_pct", highest_tier or float(g[0]))
        elif kind == "damage_reduction":
            out["damage_reduction_scope"] = g[0].lower()
            out["damage_reduction_pct"] = float(g[1])
        elif kind == "max_stat":
            out[f"max_{g[0].lower()}_pct"] = float(g[1])
        elif kind == "damage_negation":
            out[f"negation_{g[0].lower()}_pct"] = float(g[1])
        elif kind == "restore_pct":
            out[f"restore_{g[1].lower()}_pct"] = float(g[0])
        elif kind == "consumption_reduction":
            out[f"{g[0].lower()}_cost_reduction_pct"] = float(g[1])
        elif kind == "crit_mult":
            out["crit_mult_pct"] = float(g[0])
        elif kind == "gauge_boost":
            out["gauge_boost_scope"] = g[0].lower()
            out["gauge_boost_pct"] = float(g[1])
        elif kind == "cooldown_reduction":
            out["cooldown_scope"] = g[0].lower()
            out["cooldown_reduction_pct"] = float(g[1])

    dm = _DURATION_RE.search(text)
    if dm:
        out["duration_s"] = int(dm.group(1))

    for kw, trigger in _TRIGGER_MAP.items():
        if kw in low:
            out["trigger"] = trigger
            break

    return out


# ────────────────────────────────────────────────────────────────────────
# effects sheet
# ────────────────────────────────────────────────────────────────────────

def _find_header(ws, first_col_token: str = "ID") -> int | None:
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == first_col_token:
            return i
    return None


def import_effects(wb) -> dict[str, dict]:
    ws = wb["effects"]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        eid = row[0]
        if eid is None:
            continue
        try:
            eid_i = int(eid)
        except (TypeError, ValueError):
            continue
        effect_text = row[4] or ""
        out[str(eid_i)] = {
            "category": row[5] or "",
            "effect_text": effect_text,
            "stack_self": bool(row[6]) if row[6] is not None else None,
            "stack_other": bool(row[7]) if row[7] is not None else None,
            "unobtainable": bool(row[8]) if row[8] is not None else False,
            "is_dn": bool(row[2]),
            "roll": row[11] or "",
            "notes": row[12] or "",
            "parsed": parse_effect_text(effect_text),
        }
    return out


# ────────────────────────────────────────────────────────────────────────
# items sheet → named_relics.json
# ────────────────────────────────────────────────────────────────────────

_COLOR_MAP = {"Red": "R", "Blue": "B", "Green": "G", "Yellow": "Y"}
_CHAR_TAG_RE = re.compile(r"\[([A-Za-z][A-Za-z ]+?)\]")
_KNOWN_CHARS = {
    "wylder", "guardian", "ironeye", "duchess", "raider",
    "revenant", "recluse", "executor", "undertaker", "scholar",
}


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_") or "relic"


def _detect_character(effect_names: list[str], item_name: str, location: str) -> str | None:
    # Priority: [Character] tag in one of the effect names.
    for n in effect_names:
        if not n:
            continue
        m = _CHAR_TAG_RE.search(n)
        if m:
            cid = m.group(1).strip().lower().replace(" ", "_")
            if cid in _KNOWN_CHARS:
                return cid
    # Fallback: mention in item name / location.
    haystack = f"{item_name or ''} {location or ''}".lower()
    for c in _KNOWN_CHARS:
        if c in haystack:
            return c
    return None


def _classify_source(type_raw: str | None, location: str | None,
                     item_name: str | None, detected_char: str | None) -> tuple[str, str]:
    loc = (location or "").strip()
    t = (type_raw or "").lower()
    low_loc = loc.lower()
    if "signboard" in low_loc or "sovereign sigil" in low_loc:
        return "boss_everdark", loc
    if "murk" in low_loc or "bazaar" in low_loc:
        return "shop", loc
    if t == "quest" or detected_char is not None:
        return "remembrance", loc
    if t == "boss":
        # Distinguish DLC from standard (Harmonia / Straghess / Standard-Bearers)
        low = (loc + " " + (item_name or "")).lower()
        if any(k in low for k in ("harmonia", "straghess", "standard-bearer", "dlc")):
            return "boss_dlc", loc
        return "boss_standard", loc
    if t == "shop":
        return "shop", loc
    return "other", loc


def import_items(wb) -> list[dict]:
    ws = wb["items"]
    header = _find_header(ws) or 4
    relics: list[dict] = []
    seen_ids: set[str] = set()
    for row in ws.iter_rows(min_row=header + 1, values_only=True):
        item_id = row[0]
        if item_id is None or not row[4]:
            continue
        is_dn = bool(row[1])
        color = _COLOR_MAP.get(row[2], "U") if row[2] else "U"
        type_raw = row[3]
        name = str(row[4])
        e1, etxt1 = row[5], row[6]
        e2, etxt2 = row[7], row[8]
        e3, etxt3 = row[9], row[10]
        location = row[11] or ""
        notes = row[12] or ""

        attrs: list[int] = []
        for e in (e1, e2, e3):
            if e is None:
                continue
            try:
                attrs.append(int(e))
            except (TypeError, ValueError):
                pass

        effect_names = [etxt1 or "", etxt2 or "", etxt3 or ""]
        character = _detect_character(effect_names, name, location)
        source_type, source_detail = _classify_source(type_raw, location, name, character)

        desc_parts = [n for n in effect_names if n]
        description = " · ".join(desc_parts) if desc_parts else ""

        base_id = f"{_slug(name)}_{int(item_id)}"
        rid = base_id
        # Dedupe in case of name collisions.
        i = 2
        while rid in seen_ids:
            rid = f"{base_id}_{i}"
            i += 1
        seen_ids.add(rid)

        relics.append({
            "id": rid,
            "name": name,
            "source": source_type,  # legacy key (kept for backward compat)
            "source_type": source_type,
            "source_detail": source_detail,
            "character": character or "any",
            "color": color,
            "is_dn": is_dn,
            "description": description,
            "attrs": attrs,
            "locked_labels": [],
            "attrs_verified": bool(attrs),  # true iff we have effect IDs
            "notes": notes or "",
            "xlsx_item_id": int(item_id),
        })
    return relics


# ────────────────────────────────────────────────────────────────────────
# buffs sheet → dormant powers
# ────────────────────────────────────────────────────────────────────────

def import_buffs(wb) -> list[dict]:
    ws = wb["buffs"]
    header = _find_header(ws) or 4
    out: list[dict] = []
    for row in ws.iter_rows(min_row=header + 1, values_only=True):
        bid = row[0]
        desc = row[1]
        if bid is None or not desc:
            # Variant rows (row[10] = VARIANT) reference their parent via bid.
            continue
        values: list[float] = []
        raw_vals = row[3]
        if raw_vals:
            for piece in str(raw_vals).split("/"):
                p = piece.strip()
                if not p:
                    continue
                try:
                    values.append(float(p))
                except ValueError:
                    pass
        out.append({
            "id": int(bid),
            "description": desc,
            "effect_template": row[2] or "",
            "values": values,
            "format": row[4] or "pct",
            "category": row[5] or "",
            "stack_self": bool(row[6]) if row[6] is not None else None,
            "stack_other": bool(row[7]) if row[7] is not None else None,
            "unobtainable": bool(row[8]) if row[8] is not None else False,
            "notes": row[11] or "",
        })
    return out


# ────────────────────────────────────────────────────────────────────────
# bosses sheet
# ────────────────────────────────────────────────────────────────────────

def _to_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def import_bosses(wb) -> list[dict]:
    ws = wb["bosses"]
    header = _find_header(ws) or 4
    out: list[dict] = []
    for row in ws.iter_rows(min_row=header + 1, values_only=True):
        bid, name, npc_id, type_, threat = row[0:5]
        bid_i = _to_int(bid)
        npc_id_i = _to_int(npc_id)
        if bid_i is None and npc_id_i is None:
            continue
        out.append({
            "id": bid_i,
            "npc_id": npc_id_i,
            "raw_npc_id": str(npc_id) if npc_id is not None and npc_id_i is None else None,
            "name": name,
            "type": type_,
            "threat": threat,
            "notes": row[8] or "",
        })
    return out


# ────────────────────────────────────────────────────────────────────────
# main
# ────────────────────────────────────────────────────────────────────────

def _backup(path: Path) -> None:
    if not path.exists():
        return
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    bak = path.with_suffix(path.suffix + f".bak.{stamp}")
    shutil.copy2(path, bak)
    print(f"  backup → {bak.name}")


def main() -> int:
    cli = argparse.ArgumentParser(
        description=(
            "Regenerate data/{effects_enriched,named_relics,buffs,bosses}.json "
            "from the relics.pro compendium xlsx. Only needed when a game "
            "patch introduces new effects; the committed JSONs already hold "
            "the last-seen snapshot."
        ),
    )
    cli.add_argument("--xlsx", type=Path, required=True,
                     help="Path to the relics.pro compendium .xlsx file")
    args = cli.parse_args()
    if not args.xlsx.exists():
        print(f"ERROR: xlsx not found at {args.xlsx}")
        return 1
    wb = load_workbook(args.xlsx, data_only=True)
    print(f"Loaded {args.xlsx}")
    print(f"Sheets: {wb.sheetnames}\n")

    effects = import_effects(wb)
    with_prose = sum(1 for v in effects.values() if v["effect_text"])
    with_parsed = sum(1 for v in effects.values() if v["parsed"])
    out = DATA_DIR / "effects_enriched.json"
    _backup(out)
    out.write_text(json.dumps({
        "_source": "relics.pro compendium · nightreign data.xlsx",
        "_schema": "effect_id (string) -> { category, effect_text, stack_self, stack_other, unobtainable, is_dn, roll, notes, parsed:{damage_mult_pct?, damage_scope?, max_hp_pct?, negation_*_pct?, restore_*_pct?, duration_s?, trigger?, ...} }",
        "effects": effects,
    }, indent=2, ensure_ascii=False))
    print(f"✓ effects_enriched.json  {len(effects)} total, {with_prose} with prose, {with_parsed} auto-parsed")

    relics = import_items(wb)
    verified = sum(1 for r in relics if r["attrs_verified"])
    out = DATA_DIR / "named_relics.json"
    _backup(out)
    out.write_text(json.dumps({
        "_source": "relics.pro compendium · items sheet (auto-imported)",
        "_schema": "Named/unique relics with explicit effect IDs. E1/E2/E3 columns from xlsx are mapped to attrs[]. source_type ∈ {remembrance, shop, boss_standard, boss_everdark, boss_dlc, other}.",
        "_colors_legend": {
            "R": "Red (Burning Scene)",
            "G": "Green (Tranquil Scene)",
            "B": "Blue (Drizzly Scene)",
            "Y": "Yellow (Luminous Scene)",
            "U": "Universal",
        },
        "_source_types": {
            "remembrance": "Character-specific Remembrance quest reward",
            "shop": "Small Jar Bazaar (Murk) or Collector Signboard",
            "boss_standard": "First-kill drop from a base Nightlord",
            "boss_everdark": "Collector Signboard · 12 Sovereign Sigils",
            "boss_dlc": "Forsaken Hollows DLC boss drop",
            "other": "Tutorial / random expedition reward / special event",
        },
        "relics": relics,
    }, indent=2, ensure_ascii=False))
    print(f"✓ named_relics.json  {len(relics)} relics, {verified} verified")

    buffs = import_buffs(wb)
    out = DATA_DIR / "buffs.json"
    _backup(out)
    out.write_text(json.dumps({
        "_source": "relics.pro compendium · buffs sheet",
        "_schema": "Dormant Powers (permanent buffs earned per expedition). values[] holds the tier levels (typically tier 1 / tier 2).",
        "buffs": buffs,
    }, indent=2, ensure_ascii=False))
    print(f"✓ buffs.json  {len(buffs)} dormant powers")

    bosses = import_bosses(wb)
    out = DATA_DIR / "bosses.json"
    _backup(out)
    out.write_text(json.dumps({
        "_source": "relics.pro compendium · bosses sheet",
        "bosses": bosses,
    }, indent=2, ensure_ascii=False))
    print(f"✓ bosses.json  {len(bosses)} bosses")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
